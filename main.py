import hashTable
import openpyxl
from pathlib import Path
from itertools import combinations



dt = Path("C:/Users/josep/Documents/C950/WGUPS Distance Table.xlsx")
dtWb = openpyxl.load_workbook(dt, data_only=True)
dtSheet = dtWb.active
col = 9
rw = 2
distances = hashTable.hashTable(200)
i = 1
print(dtSheet[9][2].value)
while i < 27:
    dist = dtSheet[col][rw].value
    if dist == 0:
        col += 1
        i += 1
        rw = 2
        continue
    loc1 = dtSheet[col][0].value
    loc2 = dtSheet[8][rw].value
    distances.set(loc1 + ' & ' + loc2, dist)
    rw += 1
print(distances.get(loc1 + ' & ' + loc2))




pf = Path("C:/Users/josep/Documents/C950/WGUPS Package File.xlsx")
pfWb = openpyxl.load_workbook(pf)
pfSheet = pfWb.active
packageList = []
sublist = []
for row in pfSheet.iter_rows():
    if len(sublist) > 0:
        packageList.append(sublist.copy())
    sublist.clear()
    if type(row[0].value) != int:
        continue
    for cell in row:

        sublist.append(cell.value)
length = len(packageList)
packages = hashTable.hashTable(length)

for row in packageList:
    key = row.pop(0)
    packages.set(key, [row])
print(packages.get(1))
